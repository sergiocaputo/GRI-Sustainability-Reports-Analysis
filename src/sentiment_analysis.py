import json
import os
import shutil
import openpyxl
import requests

from langdetect import detect
from textblob import TextBlob
from GRIs import GRI
from excel_mapping import excel_map


def compute_sentiment(GRI_dir):
    """
    Compute Sentiment Analysis polarity and subjectivity on each context for each GRI disclosure found in the reports
    :param GRI_dir: Path to the directory containing the json files about GRI disclosures in reports
    """
    for doc in os.listdir(GRI_dir):
        doc_file = os.path.join(GRI_dir, doc)
        print(doc_file)
       
        with open(doc_file) as f:
            json_GRI = json.load(f)

        for k in GRI.keys():
            if json_GRI[k] != False:
                context = json_GRI[k]['context_before']+ " " + json_GRI[k]['context_after']
                try:
                    sentiment_context_language = detect(context)

                    if sentiment_context_language == 'en':
                        sentiment_context = TextBlob(context)
                        json_GRI[k]['context_sentiment'] = map_sentiment(sentiment_context.sentiment.polarity)
                        json_GRI[k]['context_subjectivity'] = map_subjectivity(sentiment_context.sentiment.subjectivity)
                    
                    else:
                        url = "http://193.204.187.210:9009/sentipolc/v1/classify"
                        payload =  '{ "texts": [{"id": "1","text": "'+ context +'"}]}'
                        headers = {
                                    'Content-Type': 'text/plain'
                                }

                        response = requests.request("POST", url, headers=headers, data=payload)
                        response = json.loads(response.text)

                        json_GRI[k]['context_sentiment'] = response["results"][0]["polarity"]
                        json_GRI[k]['context_subjectivity'] = response["results"][0]["subjectivity"]
                    
                except:
                    json_GRI[k]['context_sentiment'] = 'neutral'
                    json_GRI[k]['context_subjectivity'] = 'obj'
                    
                with open(doc_file, 'w') as f:
                    json.dump(json_GRI, f, indent = 4)


def GRI_reports_to_excel_with_sentiment(GRI_excel_file, GRI_dir):
    """
    Save GRI disclosure presences in Excel considering also the context's sentiment of disclosures (neutral or positive sentiment only)
    :param GRI_excel_file: Path to the directory where excel file will be stored
    :param GRI_dir: Path to the directory where GRI json are saved
    """
    GRI_excel_file = os.path.join('excel', GRI_excel_file)
    shutil.copyfile('excel\evaluation_GRI_template.xlsx', GRI_excel_file)
    
    excel_file = openpyxl.load_workbook(GRI_excel_file)
    excel_worksheet = excel_file.worksheets[0]

    for doc in range(1, len(os.listdir(GRI_dir)) + 1):
        doc_file = os.path.join(GRI_dir, str(doc) + "_GRI_report.json")

        with open(doc_file) as f:
            json_GRI = json.load(f)

        for k in excel_map.keys():
            # for disclosures without subcategories (ex. GRI 200, GRI 300...)
            if k.find("-") == -1:
                if json_GRI["GRI " + k] != False and json_GRI["GRI " + k]["context_sentiment"] != "neg":
                    excel_worksheet.cell(doc *3, excel_map[k] + 1).value = json_GRI["GRI " + k]["page"]
                else:
                    excel_worksheet.cell(doc *3, excel_map[k] + 1).value = 'no'
            else:
                # disclosures about subcategories (ex. 201-1, 201-2, 201-3...)
                if json_GRI[k] != False and json_GRI[k]["context_sentiment"] != "neg":
                    excel_worksheet.cell(doc *3, excel_map[k] + 1).value = json_GRI[k]["page"]
                else:
                    # disclosures about subcategories preceded by GRI (ex. GRI 201-1, GRI 201-2, GRI 201-3...)
                    if json_GRI["GRI " + k] != False and json_GRI["GRI " + k]["context_sentiment"] != "neg":
                        excel_worksheet.cell(doc *3, excel_map[k] + 1).value = json_GRI["GRI " + k]["page"]
                    else:
                        excel_worksheet.cell(doc *3, excel_map[k] + 1).value = 'no'

    excel_file.save(GRI_excel_file)
    excel_file.close


def map_sentiment(val):
    if val > 0.5:
        return 'pos'
    elif val < -0.5:
        return 'neg'
    else:
        return 'neutral'


def map_subjectivity(val):
    if val > 0.5:
        return 'subj'
    else:
        return 'obj'
    