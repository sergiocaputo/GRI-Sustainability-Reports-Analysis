import json
import re
import shutil
import openpyxl
import os
import pdfplumber
from pytesseract import image_to_string
from pdf2image import convert_from_path
from excel_mapping import excel_map


def convert_pdf_to_img(pdf_path):
    return convert_from_path(pdf_path, poppler_path=r'C:\Program Files\poppler-22.01.0\Library\bin')


def convert_img_to_text(img):
    return image_to_string(img)


def text_from_pdf(pdf_path, texts_path):
    images = convert_pdf_to_img(pdf_path)
    i = 0
    for imm in images:
        imm.save(os.path.join(texts_path, str(i) + ".jpg"), 'JPEG')
        with open(os.path.join(texts_path, str(i) + ".txt"), 'w') as f:
            f.write(convert_img_to_text(imm))
        i = i + 1


def get_text_from_pdf(pdf_path, texts_path):
    with pdfplumber.open(pdf_path) as pdf:
        i = 0
        for page in pdf.pages:
            with open(os.path.join(texts_path, str(i) + ".txt"), 'w', encoding="utf-8") as f:
                f.write(page.extract_text())
                f.close()
            i = i + 1
    pdf.close()


def GRI_reports_to_excel(GRI_excel_file, GRI_dir):
    """
    Save GRI disclosure presences in Excel 
    :param GRI_excel_file: Path to the directory where excel file will be stored
    :param GRI_dir: Path to the directory where GRI json are saved
    """
    GRI_excel_file = os.path.join('excel', GRI_excel_file)
    shutil.copyfile('excel\evaluation_GRI_template.xlsx', GRI_excel_file)
    
    excel_file = openpyxl.load_workbook(GRI_excel_file)
    excel_worksheet = excel_file.worksheets[0]

    for doc in range(1, len(os.listdir(GRI_dir)) + 1):
        doc_file = os.path.join(GRI_dir, str(doc) + "_GRI_report.json")

        with open(doc_file) as f:
            json_GRI = json.load(f)

        for k in excel_map.keys():
            # for disclosures without subcategories (ex. GRI 200, GRI 300...)
            if k.find("-") == -1:
                if json_GRI["GRI " + k] != False:
                    excel_worksheet.cell(doc *3, excel_map[k] + 1).value = json_GRI["GRI " + k]["page"]
                else:
                    excel_worksheet.cell(doc *3, excel_map[k] + 1).value = 'no'
            else:
                # disclosures about subcategories (ex. 201-1, 201-2, 201-3...)
                if json_GRI[k] != False:
                    excel_worksheet.cell(doc *3, excel_map[k] + 1).value = json_GRI[k]["page"]
                else:
                    # disclosures about subcategories preceded by GRI (ex. GRI 201-1, GRI 201-2, GRI 201-3...)
                    if json_GRI["GRI " + k] != False:
                        excel_worksheet.cell(doc *3, excel_map[k] + 1).value = json_GRI["GRI " + k]["page"]
                    else:
                        excel_worksheet.cell(doc *3, excel_map[k] + 1).value = 'no'

    excel_file.save(GRI_excel_file)
    excel_file.close


def compute_metrics(GRI_excel_file, GRI_dir):
    GRI_excel_file = os.path.join('excel', GRI_excel_file)

    excel_file = openpyxl.load_workbook(GRI_excel_file)
    excel_worksheet = excel_file.worksheets[0]

    for doc in range(1, len(os.listdir(GRI_dir)) + 1):
        r = doc *3 - 1

        for discolsure in [9,42]:
            # 9th column in excel: starting from this column we can compute metrics considering general disclosures + specific disclosures
            # 42nd column in excel: starting from this column we can compute metrics considering only specific disclosures
            correct_no = 0
            correct_yes = 0
            false_negative = 0
            false_positive = 0

            for c in range(discolsure, 127):
                if excel_worksheet.cell(r,c).value == excel_worksheet.cell(r + 1, c).value:
                    correct_no = correct_no + 1
                else:
                    if excel_worksheet.cell(r,c).value == 'yes' and excel_worksheet.cell(r + 1, c).value.find('.txt') and excel_worksheet.cell(r + 1, c).value != 'no':
                        correct_yes = correct_yes + 1
                    else: 
                        if excel_worksheet.cell(r,c).value == 'yes' and excel_worksheet.cell(r + 1,c).value == 'no':
                            false_negative = false_negative + 1
                        else:
                            false_positive = false_positive + 1

            if discolsure == 9:
                displacement = 0
            else:
                displacement = 9

            excel_worksheet.cell(r + 1, 129 + displacement).value = correct_yes
            excel_worksheet.cell(r + 1, 130 + displacement).value = correct_no
            excel_worksheet.cell(r + 1, 131 + displacement).value = false_negative
            excel_worksheet.cell(r + 1, 132 + displacement).value = false_positive
            
            # precision
            if (correct_yes + false_positive) != 0:
                excel_worksheet.cell(r + 1, 133 + displacement).value = correct_yes/(correct_yes + false_positive)
            else:
                excel_worksheet.cell(r + 1, 133 + displacement).value = "N/A"

            # recall
            if (correct_yes + false_negative) != 0:
                excel_worksheet.cell(r + 1, 134 + displacement).value = correct_yes/(correct_yes + false_negative)
            else:
                excel_worksheet.cell(r + 1, 134 + displacement).value = "N/A"
            
    excel_file.save(GRI_excel_file)
    excel_file.close
        

def merge_excels(excel_file_ocr, excel_file_extract, excel_file_join):
    """
    Combine the results of method with OCR and with extractor 
    :param excel_file_ocr: Excel file with OCR results
    :param excel_file_extract: Excel file with extractor results
    :param excel_file_join: Excel file that will be produced with the join of ocr and extract results
    """
    excel_file_ocr = os.path.join('excel', excel_file_ocr)
    excel_file_extract = os.path.join('excel', excel_file_extract)
    excel_file_join = os.path.join('excel', excel_file_join)

    shutil.copyfile('excel\evaluation_GRI_template.xlsx', excel_file_join)

    excel_file = openpyxl.load_workbook(excel_file_join)
    excel_worksheet_join = excel_file.worksheets[0]

    excel_worksheet_ocr = openpyxl.load_workbook(excel_file_ocr).worksheets[0]
    excel_worksheet_extract = openpyxl.load_workbook(excel_file_extract).worksheets[0]

    for doc in range(1, len(os.listdir('GRI_reports_ocr')) + 1):
        r = doc *3

        for c in range(9, 127):
            # equal case: both OCR and extractor found the same GRI reference in the same page
            if excel_worksheet_ocr.cell(r,c).value == excel_worksheet_extract.cell(r, c).value:
                excel_worksheet_join.cell(r,c).value = excel_worksheet_ocr.cell(r,c).value
            else:
                # one has page, one not: only one method found the GRI reference
                if excel_worksheet_ocr.cell(r,c).value == 'no' or excel_worksheet_extract.cell(r,c).value == 'no':
                    if excel_worksheet_ocr.cell(r,c).value != 'no':
                        excel_worksheet_join.cell(r,c).value = excel_worksheet_ocr.cell(r,c).value    
                    else:
                        excel_worksheet_join.cell(r,c).value = excel_worksheet_extract.cell(r, c).value
                else:
                # two different pages: both methods found different pages for same GRI reference -> chosen the one with the greatest number of page (GRI usually at the end of docs)
                    if int(excel_worksheet_ocr.cell(r,c).value.split('.')[0]) > int(excel_worksheet_extract.cell(r,c).value.split('.')[0]):
                        excel_worksheet_join.cell(r,c).value = excel_worksheet_ocr.cell(r,c).value
                    else:
                        excel_worksheet_join.cell(r,c).value = excel_worksheet_extract.cell(r,c).value
            #print(excel_worksheet_ocr.cell(1,c).value, "ocr: " ,excel_worksheet_ocr.cell(r,c).value, "extr: ", excel_worksheet_extract.cell(r, c).value)

    excel_file.save(excel_file_join)
    excel_file.close
            

def preprocess_text(text):
    text = re.sub(r'\(cid\:(\d+)\)', "", text).strip()
    text = text.encode("ascii", "ignore").decode()
    return text
    
